import openpyxl
wb = openpyxl.load_workbook('C:\\Users\\u\\Documents\\My Tableau Repository\\Datasources\\TradeClients.xlsx')

ws = wb['Clients']
data_rows = []

#for row in ws['A1':'T1000']:
count_row_ss=0
for row in ws:
    count_row_ss = count_row_ss + 1
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)
    
print ( count_row_ss  )
print ( len(data_rows) )
for row in data_rows:
    if 'Wholesaler' in row:
        print (row[0:3])
        

ws = wb['Sheet1']
data_rows = []

#for row in ws['A1':'T1000']:
count_row_ss=0
for row in ws:
    count_row_ss = count_row_ss + 1
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)
    
print ( count_row_ss  )
print ( len(data_rows) )
for row in data_rows:
        print (row[0:3])